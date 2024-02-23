import string
from tkinter.filedialog import SaveFileDialog
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]


for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    price = cell.value * 0.9  # type: ignore
    pcell = sheet.cell(row, 4)
    pcell.value = price

val = Reference(sheet, min_row=2, max_row=4, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(val)
sheet.add_chart(chart, "e2")

wb.save("transactions2.xlsx")
